#!/usr/bin/env python3


from openpyxl import load_workbook

wb = load_workbook('balances_template.xlsx')

print(wb.sheetnames)