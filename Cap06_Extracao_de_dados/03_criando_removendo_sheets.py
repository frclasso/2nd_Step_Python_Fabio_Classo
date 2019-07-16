#!/usr/bin/env python3

import openpyxl

wb = openpyxl.Workbook()  # carrega a planilha na memoria
#print(wb.sheetnames)

wb.create_sheet()  # cria o objeto - folhas /sheets

wb.create_sheet(index=0, title='1nd Sheet')
wb.create_sheet(index=1, title='2nd Sheet')
wb.create_sheet(index=2, title='3rd Sheet')
wb.create_sheet(index=3, title='4th Sheet')

# Removendo

del wb['Sheet']
del wb['Sheet1']


print(wb.sheetnames)

wb.save('multiplas_sheets.xlsx')