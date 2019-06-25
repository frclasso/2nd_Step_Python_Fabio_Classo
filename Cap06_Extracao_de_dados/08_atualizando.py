#!/usr/bin/env python3

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')

#print(wb.sheetnames)
sheet = wb['Sheet']

prices_updates = {'Garlic':3.07,
                  'Celery':1.19,
                  'Lemon':1.15}

for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in prices_updates:
        sheet.cell(row=rowNum, column=2).value = prices_updates[produceName]


wb.save('produceSales_updated.xlsx')
print('Feito...')