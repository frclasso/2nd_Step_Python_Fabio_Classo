#!/usr/bin/env python3

import openpyxl


wb = openpyxl.load_workbook('merged.xlsx')

#print(wb.sheetnames)
sheet = wb.active
sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')

wb.save('unmerged.xlsx')

print('Feito...')