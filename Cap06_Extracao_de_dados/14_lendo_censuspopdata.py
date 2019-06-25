#!/usr/bin/env python3

import openpyxl, pprint


wb = openpyxl.load_workbook('censuspopdata.xlsx')

#print(wb.sheetnames)
sheet = wb['Population by Census Tract']
countyData = {}

# lendo linhas
for row in range(2, sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {'tracks':0, 'pop':0})
    countyData[state][county]['tracks'] += 1
    countyData[state][county]['pop'] += int(pop)

print('Escrevendo resultados')
resultFile = open('census2010.py', 'w')
resultFile.write('allData = '+ pprint.pformat(countyData))
resultFile.close()

print('Feito...')
