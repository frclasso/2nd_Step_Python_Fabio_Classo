#!/usr/bin/env python3

import openpyxl

wb = openpyxl.load_workbook('linguagens.xlsx')
print(wb.sheetnames)

sheet = wb['Sheet']

print('Qtd máxima de linhas: ', sheet.max_row)
print('Qtd máxima de colunas: ', sheet.max_column)