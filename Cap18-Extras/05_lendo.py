#!/usr/bin/env python3

import openpyxl

wb = openpyxl.load_workbook('linguagens.xlsx')
#print(wb.sheetnames)

sheet = wb['Sheet']
# obtebdo celulas
print(sheet.cell(row=1, column=1))
print(sheet.cell(row=1, column=2))

#print(sheet['A1'])

# Obtendo valores
print(sheet['A1'].value)
print(sheet['B1'].value)

print(sheet.cell(row=1, column=2).value)
print()

# Lendo colunas
for i in range(1, 6):
    print(i, sheet.cell(row=i, column=1).value)