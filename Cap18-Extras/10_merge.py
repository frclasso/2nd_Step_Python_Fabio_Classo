#!/usr/bin/env python3

import openpyxl


wb = openpyxl.Workbook()

#print(wb.sheetnames)
sheet = wb.active

sheet['A1'] = '20 celulas unidas'
sheet.merge_cells('A1:D3')

sheet['C5'] = '2 celulas unificadas'
sheet.merge_cells('C5:D5')

wb.save('merged.xlsx')

print('Feito...')